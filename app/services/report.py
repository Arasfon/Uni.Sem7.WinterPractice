import json
import math
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.shapes import Drawing, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.config import project_root
from app.storage.history_db import HistorySession

_font_path = project_root() / "app" / "fonts" / "EBGaramond.ttf"
pdfmetrics.registerFont(TTFont("EBGaramond", str(_font_path)))


KIND_RU = {
    "photo": "Фото",
    "video": "Видео",
    "stream": "Поток",
}

STATUS_RU = {
    "running": "выполняется",
    "stopped": "остановлено",
    "ok": "завершено",
    "error": "ошибка",
}


def _ts_str(ts: Optional[float]) -> str:
    if ts is None:
        return "-"
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _utc_iso(ts: Optional[float]) -> str:
    if ts is None:
        return ""
    return datetime.fromtimestamp(float(ts), tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _safe_json_loads(s: Optional[str]) -> dict[str, Any]:
    if not s:
        return {}
    try:
        v = json.loads(s)
        return v if isinstance(v, dict) else {}
    except Exception:
        return {}


def _fetch_processings(hs: HistorySession, start_ts: Optional[float], end_ts: Optional[float]) -> list[dict[str, Any]]:
    hs.conn.row_factory = None
    q = """
      SELECT id, kind, created_at, started_at, ended_at, source, status, params_json, result_json, error
      FROM processings
    """
    args: list[Any] = []
    where = []
    if start_ts is not None:
        where.append("created_at >= ?")
        args.append(float(start_ts))
    if end_ts is not None:
        where.append("created_at <= ?")
        args.append(float(end_ts))
    if where:
        q += " WHERE " + " AND ".join(where)
    q += " ORDER BY created_at ASC"

    rows = hs.conn.execute(q, args).fetchall()
    out: list[dict[str, Any]] = []
    for r in rows:
        (
            pid,
            kind,
            created_at,
            started_at,
            ended_at,
            source,
            status,
            params_json,
            result_json,
            error,
        ) = r
        out.append(
            {
                "id": pid,
                "kind": kind,
                "created_at": float(created_at) if created_at is not None else None,
                "started_at": float(started_at) if started_at is not None else None,
                "ended_at": float(ended_at) if ended_at is not None else None,
                "source": source,
                "status": status,
                "params": _safe_json_loads(params_json),
                "result": _safe_json_loads(result_json),
                "error": error,
            }
        )
    return out


def _fetch_timeline(hs: HistorySession, pid: str) -> list[tuple[float, int]]:
    rows = hs.conn.execute(
        """
        SELECT t, count
        FROM timeline_points
        WHERE processing_id = ?
        ORDER BY t ASC
        """,
        (pid,),
    ).fetchall()
    return [(float(t), int(c)) for (t, c) in rows]


def _downsample(points: list[tuple[float, int]], max_points: int = 300) -> list[tuple[float, int]]:
    if len(points) <= max_points:
        return points
    step = int(math.ceil(len(points) / max_points))
    return points[::step]


def _make_line_chart(points: list[tuple[float, int]], title: str) -> Drawing:
    w = int(round(16.5 * cm))
    h = int(round(6.0 * cm))
    d = Drawing(w, h)

    d.add(String(0, h - 12, title, fontSize=12, fontName="EBGaramond"))

    if not points:
        d.add(String(0, h // 2, "Нет точек графика", fontSize=11, fontName="EBGaramond"))
        return d

    lp = LinePlot()

    lp.x = int(round(0.7 * cm))
    lp.y = int(round(0.6 * cm))
    lp.width = int(round(w - (1.2 * cm)))
    lp.height = int(round(h - (2.0 * cm)))

    lp.data = [points]

    lp.xValueAxis.labelTextFormat = "%.0f"
    lp.yValueAxis.labelTextFormat = "%d"
    lp.xValueAxis.visibleGrid = True
    lp.yValueAxis.visibleGrid = True
    lp.xValueAxis.gridStrokeColor = colors.lightgrey
    lp.yValueAxis.gridStrokeColor = colors.lightgrey

    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    lp.xValueAxis.valueMin = float(min(xs))
    xmax = float(max(xs))
    xmin = float(min(xs))
    lp.xValueAxis.valueMax = xmax if xmax > xmin else xmin + 1.0
    lp.yValueAxis.valueMin = 0
    lp.yValueAxis.valueMax = int(max(ys) + 1)

    d.add(lp)
    return d


def _cols(conn, table: str) -> list[str]:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return [r[1] for r in rows]


def _style_header(ws, header_row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="111827")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(ws, max_width: int = 60) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            s = str(v)
            widths[i] = max(widths.get(i, 0), min(len(s) + 2, max_width))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_pdf_report(start_ts: Optional[float], end_ts: Optional[float]) -> bytes:
    hs = HistorySession.open()
    try:
        items = _fetch_processings(hs, start_ts, end_ts)
    finally:
        pass

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.6 * cm,
        bottomMargin=1.6 * cm,
        title="Отчёт счётчика велосипедов",
    )
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "EBGaramond"
    styles["Title"].fontSize = 22
    styles["Heading2"].fontName = "EBGaramond"
    styles["Heading2"].fontSize = 16
    styles["Heading3"].fontName = "EBGaramond"
    styles["Heading3"].fontSize = 13
    styles["Normal"].fontName = "EBGaramond"
    styles["Normal"].fontSize = 11
    story = []

    story.append(Paragraph("Отчёт счётчика велосипедов", styles["Title"]))
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            f"Сформирован: {_ts_str(datetime.now(tz=timezone.utc).timestamp())}<br/>"
            f"Диапазон: {start_ts or 'ВСЕ'} → {end_ts or 'ВСЕ'}<br/>"
            f"Всего обработок: {len(items)}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 10))

    by_kind: dict[str, int] = {}
    by_status: dict[str, int] = {}
    for it in items:
        by_kind[it["kind"]] = by_kind.get(it["kind"], 0) + 1
        by_status[it["status"]] = by_status.get(it["status"], 0) + 1

    summary_data = [
        ["Сводка", ""],
        ["По типу", ", ".join([f"{KIND_RU.get(k, k)}: {v}" for k, v in sorted(by_kind.items())]) or "-"],
        ["По статусу", ", ".join([f"{STATUS_RU.get(k, k)}: {v}" for k, v in sorted(by_status.items())]) or "-"],
    ]
    t = Table(summary_data, colWidths=[3.0 * cm, 13.5 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.black),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "EBGaramond"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 14))

    list_rows = [["Создано", "Тип", "Статус", "Источник", "ID"]]
    for it in items[-200:]:
        list_rows.append(
            [
                _ts_str(it["created_at"]),
                KIND_RU.get(it["kind"], it["kind"]),
                STATUS_RU.get(it["status"], it["status"]),
                it["source"] or "-",
                it["id"][:10],
            ]  # type: ignore
        )
    lt = Table(list_rows, colWidths=[4.2 * cm, 1.8 * cm, 2.0 * cm, 6.0 * cm, 2.5 * cm])
    lt.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.black),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "EBGaramond"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(Paragraph("Последние обработки (до 200)", styles["Heading3"]))
    story.append(lt)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Детали", styles["Heading2"]))
    story.append(Spacer(1, 8))

    try:
        for it in items:
            pid = it["id"]
            kind = it["kind"]
            status = it["status"]
            source = it["source"] or "-"
            created_at = _ts_str(it["created_at"])
            started_at = _ts_str(it["started_at"])
            ended_at = _ts_str(it["ended_at"])
            result = it["result"]
            err = it["error"]

            story.append(
                Paragraph(f"{KIND_RU.get(kind, kind).upper()} · {STATUS_RU.get(status, status)} · {created_at}", styles["Heading3"])  # type: ignore
            )
            story.append(
                Paragraph(
                    f"<b>ID:</b> {pid}<br/>"
                    f"<b>Источник:</b> {source}<br/>"
                    f"<b>Начало:</b> {started_at}<br/>"
                    f"<b>Конец:</b> {ended_at}",
                    styles["Normal"],
                )
            )
            story.append(Spacer(1, 4))

            if kind == "photo":
                story.append(
                    Paragraph(
                        f"<b>Количество:</b> {result.get('count', '-')}, "
                        f"<b>Рамок:</b> {result.get('boxes_count', '-')}, "
                        f"<b>Размер:</b> {result.get('width', '?')}x{result.get('height', '?')}",
                        styles["Normal"],
                    )
                )
            else:
                story.append(
                    Paragraph(
                        f"<b>Кадров:</b> {result.get('frames_out', result.get('frames_processed', '-'))} · "
                        f"<b>Среднее:</b> {result.get('avg_count', '-') } · "
                        f"<b>Макс:</b> {result.get('max_count', '-') } · "
                        f"<b>Последнее:</b> {result.get('last_count', '-') }",
                        styles["Normal"],
                    )
                )

            if err:
                story.append(Spacer(1, 3))
                story.append(Paragraph(f"<b>Ошибка:</b> {err}", styles["Normal"]))

            if kind in ("video", "stream"):
                pts = _fetch_timeline(hs, pid)
                pts = _downsample(pts, max_points=300)
                story.append(Spacer(1, 6))
                story.append(_make_line_chart(pts, title=f"{KIND_RU.get(kind, kind)} график (время / количество)"))

            story.append(Spacer(1, 12))

    finally:
        hs.close()

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()

    return pdf_bytes


def generate_xlsx_report(
    start_ts: Optional[float],
    end_ts: Optional[float],
    start_str: Optional[str],
    end_str: Optional[str],
) -> tuple[bytes, str]:
    hs = HistorySession.open()
    conn = hs.conn

    try:
        p_cols = _cols(conn, "processings")
        p_select = ", ".join([f"p.{c}" for c in p_cols])

        where = []
        args: list[Any] = []
        if start_ts is not None:
            where.append("p.created_at >= ?")
            args.append(float(start_ts))
        if end_ts is not None:
            where.append("p.created_at <= ?")
            args.append(float(end_ts))

        p_query = f"SELECT {p_select} FROM processings p"
        if where:
            p_query += " WHERE " + " AND ".join(where)
        p_query += " ORDER BY p.created_at ASC"

        p_rows = conn.execute(p_query, args).fetchall()

        tp_cols = _cols(conn, "timeline_points")

        if "processing_id" not in tp_cols:
            raise ValueError("timeline_points table missing processing_id column")

        tp_select = ", ".join([f"tp.{c}" for c in tp_cols])
        tp_query = f"""
            SELECT
              {tp_select},
              p.kind AS p_kind,
              p.created_at AS p_created_at,
              p.started_at AS p_started_at
            FROM timeline_points tp
            JOIN processings p ON p.id = tp.processing_id
        """
        if where:
            tp_query += " WHERE " + " AND ".join(where)
        tp_query += " ORDER BY p.created_at ASC, tp.t ASC"

        tp_rows = conn.execute(tp_query, args).fetchall()

    finally:
        hs.close()

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore

    ws_meta = wb.create_sheet("meta")
    now_utc = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ws_meta.append(["generated_at_utc", now_utc])
    ws_meta.append(["range_start", start_str or "ALL"])
    ws_meta.append(["range_end", end_str or "ALL"])
    ws_meta.append(["processings_count", len(p_rows)])
    ws_meta.append(["timeline_points_count", len(tp_rows)])
    _autosize(ws_meta)

    ws_p = wb.create_sheet("processings")
    headers = list(p_cols)

    def add_ts_helpers(cols: list[str]) -> list[str]:
        out = cols[:]
        for k in ("created_at", "started_at", "ended_at"):
            if k in cols:
                out.insert(out.index(k) + 1, f"{k}_utc")
        return out

    p_headers = add_ts_helpers(headers)
    ws_p.append(p_headers)
    _style_header(ws_p)
    ws_p.freeze_panes = "A2"
    ws_p.auto_filter.ref = f"A1:{get_column_letter(len(p_headers))}1"

    for row in p_rows:
        row_list = list(row)
        out_row: list[Any] = []
        for i, col in enumerate(p_cols):
            v = row_list[i]
            out_row.append(v)
            if col in ("created_at", "started_at", "ended_at"):
                out_row.append(_utc_iso(v if v is not None else None))
        ws_p.append(out_row)

    _autosize(ws_p)

    ws_tp = wb.create_sheet("timeline_points")

    tp_headers = list(tp_cols) + ["p_kind", "p_created_at", "p_created_at_utc", "p_started_at", "p_started_at_utc", "abs_ts", "abs_utc"]
    ws_tp.append(tp_headers)
    _style_header(ws_tp)
    ws_tp.freeze_panes = "A2"
    ws_tp.auto_filter.ref = f"A1:{get_column_letter(len(tp_headers))}1"

    for row in tp_rows:
        base = list(row[: len(tp_cols)])
        p_kind = row[len(tp_cols) + 0]
        p_created_at = row[len(tp_cols) + 1]
        p_started_at = row[len(tp_cols) + 2]

        abs_ts = None
        abs_utc = ""
        if p_started_at is not None and "t" in tp_cols:
            t_rel = base[tp_cols.index("t")]
            if t_rel is not None:
                abs_ts = float(p_started_at) + float(t_rel)
                abs_utc = _utc_iso(abs_ts)

        out = base + [p_kind, p_created_at, _utc_iso(p_created_at), p_started_at, _utc_iso(p_started_at), abs_ts, abs_utc]
        ws_tp.append(out)

    _autosize(ws_tp)

    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    buf.close()

    fname = "bicycle_counter_report"
    if start_str or end_str:
        fname += f"_{(start_str or 'ALL').replace(':','-')}_to_{(end_str or 'ALL').replace(':','-')}"
    fname += ".xlsx"

    return xlsx_bytes, fname
